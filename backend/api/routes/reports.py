from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from datetime import date, timedelta
from io import BytesIO
from sqlalchemy import func
from sqlalchemy.orm import Session
from backend.models.database import get_session, Product, PriceRecord, AlertRecord, AlertConfig
from backend.services.operation_logger import OperationLogger

router = APIRouter(prefix="/api/v1/reports", tags=["报表生成"])

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from backend.services.chart_generator import (
        line_chart_to_bytes,
        pie_chart_to_bytes,
        bar_chart_to_bytes
    )
    HAS_CHART_GENERATOR = True
except ImportError:
    HAS_CHART_GENERATOR = False


def format_date(dt):
    """统一日期格式为 yyyy/mm/dd"""
    if isinstance(dt, str):
        # 尝试解析字符串日期
        try:
            dt = date.fromisoformat(dt[:10])
        except:
            return dt
    if hasattr(dt, 'strftime'):
        return dt.strftime('%Y/%m/%d')
    return str(dt)


def get_date_range(report_type: str, start_date=None, end_date=None):
    """获取报表日期范围，优先使用传入的自定义范围"""
    if start_date and end_date:
        return start_date, end_date

    today = date.today()
    if report_type == "weekly":
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        return week_start, week_end
    elif report_type == "monthly":
        month_start = date(today.year, today.month, 1)
        if today.month == 12:
            month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        return month_start, month_end
    return today - timedelta(days=7), today


def get_previous_period_range(report_type: str, start_date, end_date):
    """获取上一个周期的日期范围（用于环比）"""
    period_length = (end_date - start_date).days + 1
    prev_end = start_date - timedelta(days=1)
    prev_start = prev_end - timedelta(days=period_length - 1)
    return prev_start, prev_end


@router.get("/pdf")
async def generate_pdf_report(
    report_type: str = Query("weekly", enum=["weekly", "monthly"]),
    start_date: str = Query(None, description="开始日期 yyyy/mm/dd"),
    end_date: str = Query(None, description="结束日期 yyyy/mm/dd")
):
    """生成 PDF 报表"""

    def parse_date(s):
        if not s:
            return None
        try:
            return date.fromisoformat(s.replace('/', '-'))
        except:
            return None

    parsed_start = parse_date(start_date) if start_date else None
    parsed_end = parse_date(end_date) if end_date else None
    if not HAS_REPORTLAB:
        return {"error": "reportlab 未安装，请运行: pip install reportlab"}

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError as e:
        return {"error": f"reportlab 导入失败: {str(e)}"}

    # 注册中文字体
    try:
        pdfmetrics.registerFont(TTFont('SimHei', 'C:/Windows/Fonts/simhei.ttf'))
        pdfmetrics.registerFont(TTFont('Microsoft YaHei', 'C:/Windows/Fonts/msyh.ttc'))
        CHINESE_FONT = 'SimHei'
    except Exception:
        try:
            pdfmetrics.registerFont(TTFont('Microsoft YaHei', 'C:/Windows/Fonts/msyh.ttc'))
            CHINESE_FONT = 'Microsoft YaHei'
        except Exception:
            CHINESE_FONT = 'Helvetica'

    session = get_session()
    start_date, end_date = get_date_range(report_type, parsed_start, parsed_end)

    title = "周报" if report_type == "weekly" else "月报"
    filename = f"price_{report_type}_{end_date.isoformat()}.pdf"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # 标题
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=30, alignment=1, fontName=CHINESE_FONT)
    elements.append(Paragraph(f"采购价格{title}", title_style))
    elements.append(Spacer(1, 0.5 * cm))

    # 统计周期
    period_text = f"统计周期: {format_date(start_date)} 至 {format_date(end_date)}"
    period_style = ParagraphStyle('Period', parent=styles['Normal'], fontName=CHINESE_FONT, fontSize=12)
    elements.append(Paragraph(period_text, period_style))
    elements.append(Spacer(1, 0.5 * cm))

    # 统计数据
    stats = session.query(
        func.count(func.distinct(PriceRecord.product_id)).label('product_count'),
        func.count(PriceRecord.id).label('record_count'),
        func.max(PriceRecord.price).label('max_price'),
        func.min(PriceRecord.price).label('min_price'),
        func.avg(PriceRecord.price).label('avg_price')
    ).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).first()

    overview_data = [
        ["指标", "数值"],
        ["产品数量", str(stats.product_count or 0)],
        ["价格记录数", str(stats.record_count or 0)],
        ["最高价", f"CNY {stats.max_price:.2f}" if stats.max_price else "-"],
        ["最低价", f"CNY {stats.min_price:.2f}" if stats.min_price else "-"],
        ["平均价", f"CNY {stats.avg_price:.2f}" if stats.avg_price else "-"],
    ]

    def make_style(**kwargs):
        base = kwargs.get('base', 'Normal')
        s = ParagraphStyle('Custom', parent=styles[base], fontName=CHINESE_FONT, **kwargs)
        return s

    overview_table = Table(overview_data, colWidths=[5 * cm, 5 * cm])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#409eff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), CHINESE_FONT),
        ('FONTNAME', (0, 1), (-1, -1), CHINESE_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f7fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dcdfe6')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(overview_table)
    elements.append(Spacer(1, 0.5 * cm))

    # ========== 新增数据汇总（周报）/ 趋势分析（月报） ==========
    prev_start, prev_end = get_previous_period_range(report_type, start_date, end_date)

    # 统计上期数据
    prev_stats = session.query(
        func.count(func.distinct(PriceRecord.product_id)).label('product_count'),
        func.count(PriceRecord.id).label('record_count'),
        func.avg(PriceRecord.price).label('avg_price')
    ).filter(
        PriceRecord.record_date >= prev_start,
        PriceRecord.record_date <= prev_end
    ).first()

    if report_type == "weekly":
        # 周报：新增数据汇总 + 价格变动汇总
        new_products = session.query(
            func.count(func.distinct(Product.id)).label('new_count')
        ).join(PriceRecord).filter(
            PriceRecord.record_date >= start_date,
            PriceRecord.record_date <= end_date,
            Product.created_at >= start_date
        ).scalar() or 0

        new_records = session.query(func.count(PriceRecord.id)).filter(
            PriceRecord.record_date >= start_date,
            PriceRecord.record_date <= end_date
        ).scalar() or 0

        # 价格变动统计
        trend_stats = session.query(
            PriceRecord.trend,
            func.count(PriceRecord.id).label('count')
        ).filter(
            PriceRecord.record_date >= start_date,
            PriceRecord.record_date <= end_date,
            PriceRecord.trend.in_(['涨', '跌', '平'])
        ).group_by(PriceRecord.trend).all()

        trend_dict = {t.trend: t.count for t in trend_stats}

        # 异常告警统计（通过 AlertConfig 获取 alert_type）
        alert_stats = session.query(
            AlertConfig.alert_type,
            func.count(AlertRecord.id).label('count')
        ).join(AlertConfig, AlertRecord.alert_config_id == AlertConfig.id).filter(
            AlertRecord.triggered_at >= start_date,
            AlertRecord.triggered_at <= end_date
        ).group_by(AlertConfig.alert_type).all()

        # 新增数据汇总区块
        summary_data = [
            ["项目", "数值"],
            ["本期新增产品数", str(new_products)],
            ["本期价格记录数", str(new_records)],
            ["价格上涨产品数", str(trend_dict.get('涨', 0))],
            ["价格下跌产品数", str(trend_dict.get('跌', 0))],
            ["价格平稳产品数", str(trend_dict.get('平', 0))],
        ]

        summary_table = Table(summary_data, colWidths=[5 * cm, 5 * cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#67c23a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f9eb')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5 * cm))

        # 异常告警区块
        if alert_stats:
            alert_data = [["告警类型", "触发次数"]]
            for a in alert_stats:
                alert_data.append([a.alert_type or "其他", str(a.count)])

            alert_table = Table(alert_data, colWidths=[5 * cm, 5 * cm])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6a23c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fdf6ec')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(alert_table)
            elements.append(Spacer(1, 0.5 * cm))

    else:
        # 月报：环比数据 + 趋势分析
        curr_stats = stats
        product_change = ""
        price_change = ""
        record_change = ""

        if prev_stats and prev_stats.product_count:
            prod_diff = (curr_stats.product_count - prev_stats.product_count) / prev_stats.product_count * 100
            product_change = f"{prod_diff:+.1f}%"
        else:
            product_change = "N/A"

        if prev_stats and prev_stats.avg_price and curr_stats.avg_price:
            price_diff = (curr_stats.avg_price - prev_stats.avg_price) / prev_stats.avg_price * 100
            price_change = f"{price_diff:+.1f}%"
        else:
            price_change = "N/A"

        if prev_stats and prev_stats.record_count:
            rec_diff = (curr_stats.record_count - prev_stats.record_count) / prev_stats.record_count * 100
            record_change = f"{rec_diff:+.1f}%"
        else:
            record_change = "N/A"

        # 趋势分析段落
        trend_text = ""
        if curr_stats.avg_price and prev_stats and prev_stats.avg_price:
            price_diff = (curr_stats.avg_price - prev_stats.avg_price) / prev_stats.avg_price * 100
            if price_diff > 5:
                trend_text = "本期价格呈上涨趋势，较上期上涨 {:.1f}%，建议关注成本变化。".format(price_diff)
            elif price_diff < -5:
                trend_text = "本期价格呈下降趋势，较上期下降 {:.1f}%，市场可能走弱。".format(abs(price_diff))
            else:
                trend_text = "本期价格基本平稳，波动幅度在正常范围内。"

        comparison_data = [
            ["环比项目", "上期", "本期", "变化率"],
            ["产品数量", str(prev_stats.product_count or 0), str(curr_stats.product_count or 0), product_change],
            ["价格记录数", str(prev_stats.record_count or 0), str(curr_stats.record_count or 0), record_change],
            ["平均价", f"CNY {prev_stats.avg_price:.2f}" if prev_stats and prev_stats.avg_price else "-",
             f"CNY {curr_stats.avg_price:.2f}" if curr_stats and curr_stats.avg_price else "-", price_change],
        ]

        comparison_table = Table(comparison_data, colWidths=[4 * cm, 3.5 * cm, 3.5 * cm, 3 * cm])
        comparison_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#909399')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), CHINESE_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f4f4f5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dcdfe6')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(comparison_table)
        elements.append(Spacer(1, 0.5 * cm))

        if trend_text:
            trend_para = ParagraphStyle('Trend', parent=styles['Normal'], fontName=CHINESE_FONT,
                                       fontSize=11, textColor=colors.HexColor('#303133'),
                                       spaceAfter=10, leading=16)
            elements.append(Paragraph(trend_text, trend_para))
            elements.append(Spacer(1, 0.5 * cm))

    elements.append(Spacer(1, 0.5 * cm))

    # 生成图表
    if HAS_CHART_GENERATOR:
        try:
            # 涨跌排行柱状图
            subquery = session.query(
                PriceRecord.product_id,
                func.max(PriceRecord.record_date).label('max_date')
            ).group_by(PriceRecord.product_id).subquery()

            latest_prices = session.query(
                PriceRecord.product_id,
                PriceRecord.change_percent
            ).join(
                subquery,
                (PriceRecord.product_id == subquery.c.product_id) &
                (PriceRecord.record_date == subquery.c.max_date)
            ).all()

            product_ids = [lp.product_id for lp in latest_prices]
            products = {p.id: p.product_name for p in session.query(Product).filter(Product.id.in_(product_ids)).all()}

            ranking_data = [(products.get(lp.product_id, "未知")[:10], float(lp.change_percent) if lp.change_percent else 0.0) for lp in latest_prices]
            ranking_data.sort(key=lambda x: x[1], reverse=True)
            ranking_data = ranking_data[:10]

            if ranking_data:
                categories = [r[0] for r in ranking_data]
                values = [r[1] for r in ranking_data]
                bar_buf = bar_chart_to_bytes(categories, values, "产品价格涨跌排行")
                elements.append(Image(bar_buf, width=14 * cm, height=7 * cm))
                elements.append(Spacer(1, 0.5 * cm))
        except Exception as e:
            from loguru import logger
            logger.warning(f"PDF 柱状图生成失败: {e}")

        try:
            # 价格占比饼图
            distribution = session.query(
                Product.product_name,
                func.count(PriceRecord.id).label('count')
            ).join(PriceRecord).filter(
                PriceRecord.record_date >= start_date,
                PriceRecord.record_date <= end_date
            ).group_by(Product.id).order_by(func.count(PriceRecord.id).desc()).limit(8).all()

            if distribution:
                sizes = [float(d.count) for d in distribution]
                labels = [d.product_name[:12] for d in distribution]
                pie_buf = pie_chart_to_bytes(sizes, labels, "产品价格记录占比")
                elements.append(Image(pie_buf, width=12 * cm, height=8 * cm))
        except Exception as e:
            from loguru import logger
            logger.warning(f"PDF 饼图生成失败: {e}")

    elements.append(PageBreak())

    # 产品明细表 (扩展到 100 条)
    products_data = session.query(
        Product.product_name,
        func.max(PriceRecord.price).label('max_price'),
        func.min(PriceRecord.price).label('min_price'),
        func.avg(PriceRecord.price).label('avg_price'),
        func.count(PriceRecord.id).label('record_count')
    ).join(PriceRecord).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).group_by(Product.id).order_by(func.avg(PriceRecord.price).desc()).limit(100).all()

    detail_header = [["产品名称", "最高价", "最低价", "均价", "记录数"]]
    detail_data = [[
        p.product_name[:20] if p.product_name else "N/A",
        f"CNY {p.max_price:.2f}" if p.max_price else "-",
        f"CNY {p.min_price:.2f}" if p.min_price else "-",
        f"CNY {p.avg_price:.2f}" if p.avg_price else "-",
        str(p.record_count)
    ] for p in products_data]

    detail_table = Table(detail_header + detail_data, colWidths=[6 * cm, 3 * cm, 3 * cm, 3 * cm, 2 * cm])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#409eff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), CHINESE_FONT),
        ('FONTNAME', (0, 1), (-1, -1), CHINESE_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dcdfe6')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
    ]))
    elements.append(detail_table)

    doc.build(elements)
    buffer.seek(0)

    session.close()

    # 记录操作日志
    OperationLogger.log_report_generate(
        report_type=report_type,
        date_range={"start": str(start_date), "end": str(end_date)},
        format="pdf"
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel")
async def generate_excel_report(
    report_type: str = Query("weekly", enum=["weekly", "monthly"]),
    start_date: str = Query(None, description="开始日期 yyyy/mm/dd"),
    end_date: str = Query(None, description="结束日期 yyyy/mm/dd")
):
    """生成 Excel 报表（带图表）"""

    def parse_date(s):
        if not s:
            return None
        try:
            return date.fromisoformat(s.replace('/', '-'))
        except:
            return None

    parsed_start = parse_date(start_date) if start_date else None
    parsed_end = parse_date(end_date) if end_date else None

    if not HAS_OPENPYXL:
        return {"error": "openpyxl 未安装，请运行: pip install openpyxl"}

    session = get_session()
    start_date, end_date = get_date_range(report_type, parsed_start, parsed_end)

    wb = Workbook()
    buffer = BytesIO()

    # ========== 工作表0: 总览（综合信息表） ==========
    overview_ws = wb.active
    overview_ws.title = "总览"

    # 标题
    report_title = "周报" if report_type == "weekly" else "月报"
    overview_ws.merge_cells('A1:H1')
    overview_ws['A1'] = f"采购价格{report_title} - {format_date(start_date)} 至 {format_date(end_date)}"
    overview_ws['A1'].font = Font(size=16, bold=True)
    overview_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    overview_ws.row_dimensions[1].height = 30

    # 统计数据
    stats = session.query(
        func.count(func.distinct(PriceRecord.product_id)).label('product_count'),
        func.count(PriceRecord.id).label('record_count'),
        func.max(PriceRecord.price).label('max_price'),
        func.min(PriceRecord.price).label('min_price'),
        func.avg(PriceRecord.price).label('avg_price')
    ).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).first()

    overview_ws['A3'] = "统计指标"
    overview_ws['B3'] = "数值"
    overview_ws['A3'].font = Font(bold=True)
    overview_ws['B3'].font = Font(bold=True)
    overview_ws['A3'].fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
    overview_ws['B3'].fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
    overview_ws['A3'].font = Font(bold=True, color='FFFFFF')
    overview_ws['B3'].font = Font(bold=True, color='FFFFFF')

    overview_data = [
        ["产品数量", stats.product_count or 0],
        ["价格记录数", stats.record_count or 0],
        ["最高价", f"CNY {stats.max_price:.2f}" if stats.max_price else "-"],
        ["最低价", f"CNY {stats.min_price:.2f}" if stats.min_price else "-"],
        ["平均价", f"CNY {stats.avg_price:.2f}" if stats.avg_price else "-"],
    ]
    for i, (label, value) in enumerate(overview_data, start=4):
        overview_ws[f'A{i}'] = label
        overview_ws[f'B{i}'] = value

    overview_ws.column_dimensions['A'].width = 15
    overview_ws.column_dimensions['B'].width = 15

    # ========== 总表（完整字段） ==========
    summary_ws = wb.create_sheet("价格总表")

    # 完整的价格记录数据
    summary_records = session.query(PriceRecord, Product.product_name).join(Product).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).order_by(PriceRecord.record_date.desc()).all()

    # 表头
    summary_headers = ["日期", "产品名称", "规格", "品牌", "地区", "供应商", "价格", "趋势", "涨跌幅(%)", "数据来源"]
    summary_ws.append(summary_headers)

    # 设置表头样式
    for col, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据行
    for i, (record, pname) in enumerate(summary_records, start=2):
        row_data = [
            format_date(record.record_date),
            pname,
            record.specification or "-",
            record.brand or "-",
            record.region or "-",
            record.supplier or "-",
            record.price,
            record.trend or "-",
            record.change_percent if record.change_percent else 0,
            record.source or "-"
        ]
        summary_ws.append(row_data)

        # 设置日期列格式
        summary_ws.cell(row=i, column=1).number_format = '@'  # 文本格式
        # 价格格式
        summary_ws.cell(row=i, column=7).number_format = '#,##0.00'
        # 涨跌幅格式
        summary_ws.cell(row=i, column=9).number_format = '0.00'

    # 设置列宽
    col_widths = [12, 20, 18, 12, 10, 15, 12, 8, 10, 12]
    for i, width in enumerate(col_widths, start=1):
        summary_ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

    # ========== 工作表1: 产品列表 ==========
    products_ws = wb.create_sheet("产品列表")
    products = session.query(Product).filter(Product.is_active == True).all()
    products_ws.append(["ID", "产品编码", "产品名称", "分类", "单位", "数据源"])
    for p in products:
        products_ws.append([p.id, p.product_code, p.product_name, p.category, p.unit, p.source])

    products_ws.column_dimensions['A'].width = 8
    products_ws.column_dimensions['B'].width = 15
    products_ws.column_dimensions['C'].width = 25
    products_ws.column_dimensions['D'].width = 15
    products_ws.column_dimensions['E'].width = 10
    products_ws.column_dimensions['F'].width = 15

    # ========== 工作表2: 价格历史（简化版） ==========
    history_ws = wb.create_sheet("价格历史")
    history_records = session.query(PriceRecord, Product.product_name).join(Product).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).order_by(PriceRecord.record_date.desc()).all()
    history_ws.append(["日期", "产品", "价格", "趋势", "涨跌%", "来源"])
    for record, pname in history_records:
        history_ws.append([
            format_date(record.record_date),
            pname,
            record.price,
            record.trend,
            record.change_percent if record.change_percent else 0,
            record.source
        ])

    history_ws.column_dimensions['A'].width = 12
    history_ws.column_dimensions['B'].width = 25
    history_ws.column_dimensions['C'].width = 12
    history_ws.column_dimensions['D'].width = 8
    history_ws.column_dimensions['E'].width = 10
    history_ws.column_dimensions['F'].width = 15

    # ========== 工作表3: 统计汇总 ==========
    summary_stats_ws = wb.create_sheet("统计汇总")
    stats_list = session.query(
        Product.product_name,
        func.max(PriceRecord.price).label('max_price'),
        func.min(PriceRecord.price).label('min_price'),
        func.avg(PriceRecord.price).label('avg_price'),
        func.count(PriceRecord.id).label('record_count')
    ).join(PriceRecord).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).group_by(Product.id).order_by(func.avg(PriceRecord.price).desc()).all()

    summary_stats_ws.append(["产品名称", "最高价", "最低价", "均价", "记录数"])
    for s in stats_list:
        summary_stats_ws.append([
            s.product_name,
            round(s.max_price, 2) if s.max_price else 0,
            round(s.min_price, 2) if s.min_price else 0,
            round(s.avg_price, 2) if s.avg_price else 0,
            s.record_count
        ])

    summary_stats_ws.column_dimensions['A'].width = 25
    summary_stats_ws.column_dimensions['B'].width = 12
    summary_stats_ws.column_dimensions['C'].width = 12
    summary_stats_ws.column_dimensions['D'].width = 12
    summary_stats_ws.column_dimensions['E'].width = 10

    # ========== 工作表4: 涨跌排行图表 ==========
    if stats_list:
        chart_ws = wb.create_sheet("涨跌排行")
        chart_ws.append(["产品名称", "涨跌幅(%)"])

        # 获取最新价格变化率
        subquery = session.query(
            PriceRecord.product_id,
            func.max(PriceRecord.record_date).label('max_date')
        ).group_by(PriceRecord.product_id).subquery()

        latest_prices = session.query(
            PriceRecord.product_id,
            PriceRecord.change_percent
        ).join(
            subquery,
            (PriceRecord.product_id == subquery.c.product_id) &
            (PriceRecord.record_date == subquery.c.max_date)
        ).all()

        product_ids = [lp.product_id for lp in latest_prices]
        products_map = {p.id: p.product_name for p in session.query(Product).filter(Product.id.in_(product_ids)).all()}

        ranking = [(products_map.get(lp.product_id, "未知"), float(lp.change_percent) if lp.change_percent else 0.0) for lp in latest_prices]
        ranking.sort(key=lambda x: x[1], reverse=True)
        ranking = ranking[:15]

        for name, change in ranking:
            chart_ws.append([name, change])

        # 创建柱状图
        bar_chart = BarChart()
        bar_chart.type = "bar"
        bar_chart.title = "产品价格涨跌排行"
        bar_chart.style = 10
        bar_chart.y_axis.title = "产品"
        bar_chart.x_axis.title = "涨跌幅(%)"

        data = Reference(chart_ws, min_col=2, min_row=1, max_row=len(ranking) + 1)
        cats = Reference(chart_ws, min_col=1, min_row=2, max_row=len(ranking) + 1)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.height = 12
        bar_chart.width = 20

        chart_ws.add_chart(bar_chart, "D2")

        # ========== 工作表5: 价格占比饼图 ==========
        pie_ws = wb.create_sheet("价格占比")
        distribution = session.query(
            Product.product_name,
            func.count(PriceRecord.id).label('count')
        ).join(PriceRecord).filter(
            PriceRecord.record_date >= start_date,
            PriceRecord.record_date <= end_date
        ).group_by(Product.id).order_by(func.count(PriceRecord.id).desc()).limit(8).all()

        pie_ws.append(["产品名称", "记录数"])
        for d in distribution:
            pie_ws.append([d.product_name, d.count])

        pie_chart = PieChart()
        pie_chart.title = "价格记录占比"
        pie_chart.style = 10

        pie_data = Reference(pie_ws, min_col=2, min_row=1, max_row=len(distribution) + 1)
        pie_cats = Reference(pie_ws, min_col=1, min_row=2, max_row=len(distribution) + 1)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_cats)
        pie_chart.height = 12
        pie_chart.width = 15

        pie_ws.add_chart(pie_chart, "D2")

    # ========== 工作表6: 环比分析 ==========
    prev_start, prev_end = get_previous_period_range(report_type, start_date, end_date)
    comparison_ws = wb.create_sheet("环比分析")

    comparison_ws.merge_cells('A1:F1')
    comparison_ws['A1'] = f"环比分析 - {format_date(start_date)} 至 {format_date(end_date)}（上期: {format_date(prev_start)} 至 {format_date(prev_end)}）"
    comparison_ws['A1'].font = Font(size=14, bold=True)
    comparison_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    comparison_ws.row_dimensions[1].height = 25

    comparison_headers = ["产品名称", "上期均价", "本期均价", "均价变化", "记录数变化", "趋势"]
    comparison_ws.append(comparison_headers)

    for col, header in enumerate(comparison_headers, start=1):
        cell = comparison_ws.cell(row=2, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 获取上期和本期每个产品的数据
    curr_data = session.query(
        Product.id,
        Product.product_name,
        func.avg(PriceRecord.price).label('avg_price'),
        func.count(PriceRecord.id).label('record_count')
    ).join(PriceRecord).filter(
        PriceRecord.record_date >= start_date,
        PriceRecord.record_date <= end_date
    ).group_by(Product.id).all()

    curr_map = {p.id: (p.avg_price, p.record_count) for p in curr_data}

    prev_data = session.query(
        Product.id,
        func.avg(PriceRecord.price).label('avg_price'),
        func.count(PriceRecord.id).label('record_count')
    ).join(PriceRecord).filter(
        PriceRecord.record_date >= prev_start,
        PriceRecord.record_date <= prev_end
    ).group_by(Product.id).all()

    prev_map = {p.id: (p.avg_price, p.record_count) for p in prev_data}

    for product in curr_data:
        prev_info = prev_map.get(product.id)
        curr_info = curr_map.get(product.id)

        prev_avg = prev_info[0] if prev_info else None
        curr_avg = curr_info[0] if curr_info else None
        prev_cnt = prev_info[1] if prev_info else 0
        curr_cnt = curr_info[1] if curr_info else 0

        if prev_avg and curr_avg:
            price_change = f"{((curr_avg - prev_avg) / prev_avg * 100):+.1f}%"
        else:
            price_change = "N/A"

        if prev_cnt:
            count_change = f"{curr_cnt - prev_cnt:+d}"
        else:
            count_change = "+" + str(curr_cnt)

        if prev_avg and curr_avg:
            if curr_avg > prev_avg * 1.05:
                trend = "上涨"
            elif curr_avg < prev_avg * 0.95:
                trend = "下跌"
            else:
                trend = "平稳"
        else:
            trend = "新增"

        comparison_ws.append([
            product.product_name,
            f"{prev_avg:.2f}" if prev_avg else "-",
            f"{curr_avg:.2f}" if curr_avg else "-",
            price_change,
            count_change,
            trend
        ])

    comparison_ws.column_dimensions['A'].width = 25
    comparison_ws.column_dimensions['B'].width = 12
    comparison_ws.column_dimensions['C'].width = 12
    comparison_ws.column_dimensions['D'].width = 12
    comparison_ws.column_dimensions['E'].width = 12
    comparison_ws.column_dimensions['F'].width = 10

    # ========== 工作表7: 异常告警 ==========
    alert_ws = wb.create_sheet("异常告警")

    alert_ws.merge_cells('A1:E1')
    alert_ws['A1'] = f"异常告警记录 - {format_date(start_date)} 至 {format_date(end_date)}"
    alert_ws['A1'].font = Font(size=14, bold=True)
    alert_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    alert_ws.row_dimensions[1].height = 25

    # 告警汇总统计（通过 AlertConfig 获取 alert_type）
    alert_summary = session.query(
        AlertConfig.alert_type,
        func.count(AlertRecord.id).label('count')
    ).join(AlertConfig, AlertRecord.alert_config_id == AlertConfig.id).filter(
        AlertRecord.triggered_at >= start_date,
        AlertRecord.triggered_at <= end_date
    ).group_by(AlertConfig.alert_type).all()

    alert_ws.append(["告警类型", "触发次数"])
    for a in alert_summary:
        alert_ws.append([a.alert_type or "其他", a.count])

    alert_ws.cell(row=1, column=1).font = Font(bold=True)
    alert_ws.cell(row=1, column=2).font = Font(bold=True)
    alert_ws.cell(row=1, column=1).fill = PatternFill(start_color='E6A23C', end_color='E6A23C', fill_type='solid')
    alert_ws.cell(row=1, column=2).fill = PatternFill(start_color='E6A23C', end_color='E6A23C', fill_type='solid')
    alert_ws.cell(row=1, column=1).font = Font(bold=True, color='FFFFFF')
    alert_ws.cell(row=1, column=2).font = Font(bold=True, color='FFFFFF')

    alert_ws.append([])  # 空行

    # 告警明细
    alert_ws.append(["告警ID", "产品", "告警类型", "消息", "触发时间"])
    header_row = alert_ws.max_row
    for col in range(1, 6):
        cell = alert_ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='F56C6C', end_color='F56C6C', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    alert_records = session.query(
        AlertRecord, Product.product_name
    ).join(Product, AlertRecord.product_id == Product.id).filter(
        AlertRecord.triggered_at >= start_date,
        AlertRecord.triggered_at <= end_date
    ).order_by(AlertRecord.triggered_at.desc()).limit(100).all()

    for record, pname in alert_records:
        alert_ws.append([
            record.id,
            pname,
            record.alert_type or "-",
            record.alert_message or "-",
            format_date(record.triggered_at) if record.triggered_at else "-"
        ])

    alert_ws.column_dimensions['A'].width = 10
    alert_ws.column_dimensions['B'].width = 25
    alert_ws.column_dimensions['C'].width = 15
    alert_ws.column_dimensions['D'].width = 40
    alert_ws.column_dimensions['E'].width = 15
    wb.save(buffer)
    buffer.seek(0)

    filename = f"price_{report_type}_{end_date.isoformat()}.xlsx"
    session.close()

    # 记录操作日志
    OperationLogger.log_report_generate(
        report_type=report_type,
        date_range={"start": str(start_date), "end": str(end_date)},
        format="excel"
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )